####################################################################################
# G Suit Administrator for Education
#
# A G Suit for Education rendszer iskolai adminisztrációs feladatainak
# egyszerűsítésére, megkönnyítésére szolgáló alkalmazás
#
# Készítette: Venczel József
#
# Debrecen, 2020. 09. 10.
####################################################################################


#============================================================================================
# 3.1. Modulok importálása
#

import pickle
import os.path
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.errors import HttpError

import socket

# Az Excel táblázatok kezeléséhez szükséges modul
from openpyxl import load_workbook

# A felhasználónevek előállításához szükséges függvényeket tartalmazza
import unidecode

# A ebben található függvények segítségével mérem le, melyik programrész meddig fut
import time

import sys

import yaml

from random import seed
from random import random

import hashlib

#============================================================================================
# 3.2. Globális változók értékének feltöltése konfigurációs állományból
#

config = yaml.load(open("config.yaml", encoding="utf8"), Loader=yaml.FullLoader)

# A program alapértelmezés szerint <TESZT> üzemmódban működik
# Minden objektum elé odarakja a teszt változóban tárolt karaktersorozatot
# csoportnevek, Diákok és Pedagógusok szervezeti egysége
# felhasználó e-mail-ek, tanfolyam azonosítók és nevek.

if config["Tesztmód"].lower() in ["igen", "igaz", "true", "yes"]:
    teszt="teszt"
else:
    teszt=""

# A tanév kezdőévének megadása, amiből a tanév neve is származik
ev=config["Tanév kezdőéve"]
tanev=str(ev)+"-"+str(ev+1)

# A tanár és a diák felhasználókhoz társított domain (lehet a kettő ugyanaz)
suliDomain=config["Tanar e-mail címek domain része"]
diakDomain=config["Diák e-mail címek domain része"]

# Az G Suit admin felhasználóneve
adminNev=config["Admin"]

# A tanárok és titkársági dolgozók szervezeti egységének és csoportjának elnevezése
pedSZEnev=config["Pedagógusok szervezeti egységének neve"]

# A diákok szervezeti egységének és csoportjának elnevezése
diakokSZEnev=config["Diákok szervezeti egységének neve"]

# Azoknak a diákoknak a szervezeti egysége (és osztálya), akiknek nincs osztályfőnöki órája.
# Ezek lehetnek pl. cserediákok.
vendegSZEnev=config["Vendégek szervezeti egységének neve"]


#============================================================================================
# 3.3. Tanárok adatainak feldolgozása
#
# Pedagógusok és titkársági dolgozók adatainak feldolgozása
# A pedagógus belépési azonosítója meg fog egyezni a KRÉTA belépési azonosítójával.
# A tanárok esetében nem olyan egyértelmű a vezetéknév-keresztnév kettébontása.
# Nincs olyan KRÉTA export, amiben benne lennének együtt a KRÉTA felhasználónevek,
# a vezetéknév, illetve a keresztnév felbontva.
# Ezért van szükség két exportra a tanárokhoz.

print("Forrásadatok betöltése és feldolgozása...")

ped1TBL = load_workbook("AlkalmazottTablazatExport.xlsx")
ped1ML = ped1TBL.active
ped2TBL = load_workbook("Alkalmazottak_lakcimmel.xlsx")
ped2ML = ped2TBL.active

tanarEmailek = []
tanarVezNevek = []
tanarKerNevek = []
tanarOsztok = []
tanarOMek = []

for sorA, sorB in zip(ped1ML.rows, ped2ML.rows):
    if(str(sorA[5].value) and sorA[0].value != "Név"):
        tanarEmailek.append(teszt+str(sorA[5].value)+"@"+suliDomain)
        # A Google nem kezeli a titulusokat, ezért a vezetéknevekhez hozzáadjuk, ha van.
        if sorB[0].value is not None:
            tanarVezNevek.append(str(sorB[0].value)+" "+str(sorB[1].value))
        else:
            tanarVezNevek.append(str(sorB[1].value))
        tanarKerNevek.append(str(sorB[2].value))
        tanarOMek.append(str(sorA[4].value))
        tanarOsztok.append("")


#============================================================================================
# 3.4. Diákok (és kurzusok) adatainak feldolgozása
#

tbl = load_workbook("Tanulok_tantargyai_es_pedagogusai.xlsx")
ml = tbl.active

diakVezNevek = []
diakKerNevek = []
diakEmailek = []
diakOMek = []
diakOsztalya = []

kurzusNevek = []
kurzusCsopok = []
kurzusTanok = []
kurzusTanarOMek = []
kurzusAliasok = []

for sor in ml.rows:
    if str(sor[1].value) != "Vezetéknév":
        if  str(sor[3].value) not in diakOMek:
            email=unidecode.unidecode(str(sor[1].value)+str(sor[2].value)).lower()
            email=email.replace(" ", "")
            email=email+str(sor[3].value)[len(str(sor[3].value))-4:]
            diakEmailek.append(teszt+email+"@"+diakDomain)
            diakVezNevek.append(str(sor[1].value))
            diakKerNevek.append(str(sor[2].value))
            diakOMek.append(str(sor[3].value))

        # A vendégtanulókon, cserediákokon kívül mindenkinek van osztályfőnöki órája.
        # Amelyik diákhoz nem lesz osztály rendelve, azt külön csoportba és sz.e.-be tesszük.
        if str(sor[5].value)=="osztályfőnöki" :
            diakOsztalya[diakOMek.index(str(sor[3].value))] = str(sor[4].value)
            tanarOsztok[tanarOMek.index(str(sor[7].value))] = str(sor[4].value) # Ő lesz az osztályfőnök
        else:
            diakOsztalya.append("")

        kurzusNev=teszt+str(sor[4].value)+" - "+str(sor[5].value)+" ("+tanev+")"
        # Ha a kurzushoz tartozik tanár és még nincs benne a kurzusok listájában, akkor felveszi
        if sor[7].value is not None:
            if kurzusNev not in kurzusNevek:
                kurzusNevek.append(kurzusNev)
                kurzusCsopok.append(str(sor[4].value))
                kurzusTanok.append(str(sor[5].value))
                kurzusTanarOMek.append(str(sor[7].value))
                alias="d:"+teszt+str(sor[4].value)+str(sor[5].value)+tanev
                alias=alias.replace(" ", "")
                alias=alias.replace(".", "")
                alias=alias.replace("-", "")
                alias=alias.replace("/", "")
                kurzusAliasok.append(alias)
            else:
                if kurzusTanarOMek[kurzusNevek.index(kurzusNev)].find(str(sor[7].value))<0:
                    kurzusTanarOMek[kurzusNevek.index(kurzusNev)]=kurzusTanarOMek[kurzusNevek.index(kurzusNev)]+" "+str(sor[7].value)


print("Forrásadatok betöltése és feldolgozása befejeződött.")

#============================================================================================
# 3.5. Kapcsolódás a Google API szolgáltatásokhoz
#

SCOPES = ['https://www.googleapis.com/auth/admin.directory.user',
          'https://www.googleapis.com/auth/admin.directory.group',
          'https://www.googleapis.com/auth/admin.directory.orgunit',
          'https://www.googleapis.com/auth/classroom.courses',
          'https://www.googleapis.com/auth/classroom.rosters' ]

print("Kapcsolatfelvétel a szerverrel...")

creds = None
if os.path.exists('token.pickle'):
    with open('token.pickle', 'rb') as token:
        creds = pickle.load(token)

if not creds or not creds.valid:
    if creds and creds.expired and creds.refresh_token:
        creds.refresh(Request())
    else:
        flow = InstalledAppFlow.from_client_secrets_file(
            'credentials.json', SCOPES)
        creds = flow.run_local_server(port=0)

    with open('token.pickle', 'wb') as token:
        pickle.dump(creds, token)

socket.setdefaulttimeout(600)

serviceAdmin = build('admin', 'directory_v1', credentials=creds)
serviceClassroom = build('classroom', 'v1', credentials=creds)

if serviceAdmin and serviceClassroom:
    print("Szerverhez kapcsolódva.")
else:
    print("Nem sikerült kapcsolatot létesíteni a szerverrel!")


totalTime=time.time()


#============================================================================================
# 3.6. Csoport létrehozása egyszerű hibakezeléssel
#
# Vendégtanulók csoportja
#
print("Vendégtanulók csoportjának létrehozása.")
csopAz=teszt+"vendeg@"+diakDomain
try:
    csop = { "email" : csopAz,
             "description": "A vendégtanulók csoportja.",
             "name": teszt+vendegSZEnev+" osztálycsoport" }
    csop = serviceAdmin.groups().insert(body=csop).execute()
    print("A vendégtanulók csoportja létrehozva.")
except HttpError as err:
    if err.resp.status in [409]:
        print("A vendégtanulók csoportja már létezik.")
    else:
        print(err)
        sys.exit()

#============================================================================================
# 3.7. Csoport létrehozása fejlettebb hibakezeléssel
# Pedagógusok csoportja
#
print(teszt+pedSZEnev+" csoport létrehozása.")
csopAz = teszt+"ped@"+suliDomain
varido=1
ismeteld=True
while ismeteld and varido < 64:
    try:
        csop = { "email" : csopAz,
                 "description" : "A pedagógusok és titkársági dolgozók csoportja.",
                 "name" : teszt+pedSZEnev }
        csop = serviceAdmin.groups().insert(body=csop).execute()
        print("A "+teszt+pedSZEnev+" csoport létrehozva.")
        ismeteld=False
    except HttpError as err:
        if err.resp.status in [409]:
            print("A "+teszt+pedSZEnev+" csoport már létezik.")
            ismeteld=False
        elif err.resp.status in [403, 503]:
            time.sleep(varido+random())
            varido=varido*2
        else:
            print(err)
            sys.exit()
if ismeteld:
    print(err)
    sys.exit()

#============================================================================================
# 3.8. Osztálycsoportok létrehozása, kötegelt feldolgozás
#

def cb_ujOsztCsop(request_id, response, exception):

    if exception is not None and exception.resp.status not in [409]:
        
        oszt=request_id
        groupId=teszt+oszt+"@"+diakDomain
        csop = { "email" : groupId,
                 "description": "A "+oszt+" osztály csoportja.",
                 "name": teszt+oszt+" osztálycsoport" }
        ismeteld=True
        varido=1

        while ismeteld and varido < 64:
            try:
                request = serviceAdmin.groups().insert(body=csop).execute()
                ismeteld=False
            except HttpError as err:
                if err.resp.status in [409]:
                    ismeteld=False
                elif err.resp.status in [403, 503]:
                    time.sleep(varido+random())
                    varido=varido*2
                else:
                    print(err)
                    sys.exit()
            print("*", end="")

        if ismeteld:
            print(err)
            sys.exit()


startTime=time.time()
print("Osztálycsoportok létrehozása", end ="")
batch = serviceAdmin.new_batch_http_request(callback=cb_ujOsztCsop)
kotegmeret=10
db=0
for oszt in tanarOsztok:
    if oszt:
        groupId=teszt+oszt+"@"+diakDomain
        csop = { 'email' : groupId,
                 'description': 'A '+oszt+' osztály csoportja.',
                 'name': teszt+oszt+' osztálycsoport' }
        request = serviceAdmin.groups().insert(body=csop)
        batch.add(request, request_id=oszt)
        kotegmeret-=1
        print(".", end ="")
        if kotegmeret < 1:
            batch.execute()
            kotegmeret=10
            batch = serviceAdmin.new_batch_http_request(callback=cb_ujOsztCsop)
        db+=1

print(".")
batch.execute()
print(db, "osztálycsoport létrehozva.")
elapsedTime = time.time() - startTime
print("Ez a programrész %.3fs alatt futott le." % elapsedTime)


#============================================================================================
# 3.9. A szeszélyes szervezeti egységek
#
# Pedagógusok
#

print(pedSZEnev+" szervezeti egység létrehozása.")

pedSZE=serviceAdmin.orgunits().get(orgUnitPath=teszt+pedSZEnev, customerId="my_customer").execute()

if not pedSZE:
    varido=1
    ismeteld=True
    while ismeteld and varido < 64:
        try:
            pedSZE = { "name" : teszt+pedSZEnev,
                    "description" : "A tanárokat és a titkársági dolgozókat tartalmazó szervezeti egység.",
                    "parentOrgUnitPath" : "/" }
            pedSZE = serviceAdmin.orgunits().insert(body=pedSZE, customerId="my_customer").execute()
            print(pedSZEnev+" szervezeti egység létrehozva.")
            ismeteld=False
        except HttpError as err:
            if err.resp.status in [409, 400]:
                print(teszt+pedSZEnev+" szervezeti egység már létezik.")
                ismeteld=False
            elif err.resp.status in [403, 503]:
                time.sleep(varido+random())
                varido=varido*2
            else:
                print("Hiba!",err)
                sys.exit()
    if ismeteld:
        print(err)
        sys.exit()
else:
    print(teszt+pedSZEnev+" szervezeti egység már létezik.")


#
# Diákok
#

print(diakokSZEnev+" szervezeti egység létrehozása.")

diakokSZE=serviceAdmin.orgunits().get(orgUnitPath=teszt+diakokSZEnev, customerId="my_customer").execute()

if not diakokSZE:
    varido=1
    ismeteld=True
    while ismeteld and varido < 64:
        try:
            diakokSZE = {
                "name" : teszt+diakokSZEnev,
                "description" : "A diákokat, azok osztályait, tanévenként csoportosítva tartalmazó szervezeti egység.",
                "parentOrgUnitPath" : "/" }
            diakokSZE = serviceAdmin.orgunits().insert(body=diakokSZE, customerId="my_customer").execute()
            print(diakokSZEnev+" szervezeti egység létrehozva.")
            ismeteld=False
        except HttpError as err:
            if err.resp.status in [409, 400]:
                print(teszt+diakokSZEnev+" szervezeti egység már létezik.")
                ismeteld=False
            elif err.resp.status in [403, 503]:
                time.sleep(varido+random())
                varido=varido*2
            else:
                print(err)
                sys.exit()
    if ismeteld:
        print(err)
        sys.exit()
else:
    print(teszt+diakokSZEnev+" szervezeti egység már létezik.")


#
# Tanév szervezeti egysége
#

print("/"+teszt+diakokSZEnev+"/"+tanev+" szervezeti egység létrehozása.")

tanevSZE=serviceAdmin.orgunits().get(orgUnitPath=teszt+diakokSZEnev+"/"+tanev, customerId="my_customer").execute()

if not tanevSZE:
    varido=1
    ismeteld=True
    while ismeteld and varido < 64:
        try:
            tanevSZE = { "name" : tanev,
                        "description" : "Az osztályokat tartalmazó tanév szervezeti egység.",
                        "parentOrgUnitPath" : "/"+teszt+diakokSZEnev }
            tanevSZE = serviceAdmin.orgunits().insert(body=tanevSZE, customerId="my_customer").execute()
            print("A /"+teszt+diakokSZEnev+"/"+tanev+" szervezeti egység létrehozva.")
            ismeteld=False
        except HttpError as err:
            if err.resp.status in [409, 400]:
                print("A /"+teszt+diakokSZEnev+"/"+tanev+" szervezeti egység már létezik.")
                ismeteld=False
            elif err.resp.status in [403, 503]:
                time.sleep(varido+random())
                varido=varido*2
            else:
                print(err)
                sys.exit()
    if ismeteld:
        print(err)
        sys.exit()
else:
    print("A /"+teszt+diakokSZEnev+"/"+tanev+" szervezeti egység már létezik.")


#
# Osztályok szervezeti egységei a tanéven belül
#

letezo_osztalyok = []
lista = serviceAdmin.orgunits().list(customerId="my_customer", orgUnitPath="/"+teszt+diakokSZEnev+"/"+tanev).execute()
for sor in lista.get("organizationUnits"):
    letezo_osztalyok.append(str(sor.get("name")))

def cb_ujOsztSZE(request_id, response, exception):

    if exception is not None:

        oszt=request_id
        osztSZE = { "name" : oszt,
                    "description" : "A "+oszt+" osztály tanulóit tartalmazó szervezeti egység.",
                    "parentOrgUnitPath" : "/"+teszt+diakokSZEnev+"/"+tanev }
        ismeteld=True
        varido=1

        while ismeteld and varido < 64:
            try:
                request = serviceAdmin.orgunits().insert(body=osztSZE, customerId="my_customer").execute()
                ismeteld=False
            except HttpError as err:
                if err.resp.status in [409, 400]:
                    ismeteld=False
                elif err.resp.status in [403, 503]:
                    time.sleep(varido+random())
                    varido=varido*2
                else:
                    print(err)
                    sys.exit()
            print("*", end="")
        if ismeteld:
            print(err)
            sys.exit()


startTime=time.time()
print("/"+teszt+diakokSZEnev+"/"+tanev+"/osztály szervezeti egységek létrehozása", end ="")
batch = serviceAdmin.new_batch_http_request(callback=cb_ujOsztSZE)
kotegmeret=10
db=0
for oszt in tanarOsztok:
    if oszt and oszt not in letezo_osztalyok:
        osztSZE = { "name" : oszt,
                    "description" : "A "+oszt+" osztály tanulóit tartalmazó szervezeti egység.",
                    "parentOrgUnitPath" : "/"+teszt+diakokSZEnev+"/"+tanev }
        request = serviceAdmin.orgunits().insert(body=osztSZE, customerId="my_customer")
        batch.add(request, request_id=oszt)
        kotegmeret-=1
        if kotegmeret < 1:
            batch.execute()
            kotegmeret=10
            batch = serviceAdmin.new_batch_http_request(callback=cb_ujOsztSZE)
        db+=1
        print(".", end ="")

batch.execute()
print(".")
print(db, "db szervezeti egység létrehozva.")
elapsedTime = time.time() - startTime
print("Ez a programrész %.3fs alatt futott le." % elapsedTime)


#
# Vendégtanulók
#
print("Vendégtanulók szervezeti egységének létrehozása.")

vendegSZE=serviceAdmin.orgunits().get(orgUnitPath=teszt+diakokSZEnev+"/"+tanev+"/"+vendegSZEnev, customerId="my_customer").execute()

if not vendegSZE:
    varido=1
    ismeteld=True
    while ismeteld and varido < 64:
        try:
            vendegSZE = { "name" : vendegSZEnev,
                        "description" : "A vendégtanulókat tartalmazó szervezeti egység.",
                        "parentOrgUnitPath" : "/"+teszt+diakokSZEnev+"/"+tanev }
            vendegSZE = serviceAdmin.orgunits().insert(body=vendegSZE, customerId='my_customer').execute()
            print("A vendégtanulók szervezeti egysége létrehozva.")
            ismeteld=False
        except HttpError as err:
            if err.resp.status in [409, 400]:
                print("A vendégtanulók szervezeti egysége már létezik.")
                ismeteld=False
            elif err.resp.status in [403, 503]:
                time.sleep(varido+random())
                varido=varido*2
            else:
                print(err)
                sys.exit()
    if ismeteld:
        print(err)
        sys.exit()
else:
    print("A vendégtanulók szervezeti egysége már létezik.")


#============================================================================================
# 3.10. Pedagógus felhasználók létrehozása
#       (és elhelyezése a megfelelő szervezeti egységekben)
#

def cb_ujTanar(request_id, response, exception):
    if exception is not None and exception.resp.status not in [409]:
        
        om=request_id
        tanarIdx=tanarOMek.index(om)
        
        email=tanarEmailek[tanarIdx]
        veznev=tanarVezNevek[tanarIdx]
        kernev=tanarKerNevek[tanarIdx]

        tanar = {
            "primaryEmail" : email,
            "name" : {
                "givenName" : kernev,
                "familyName" : veznev
            },
            "password" : hashlib.sha1(om.encode()).hexdigest(),
            "hashFunction" : "SHA-1",
            "changePasswordAtNextLogin" : True,
            "orgUnitPath" : "/"+teszt+pedSZEnev
        }
        
        ismeteld=True
        varido=1

        while ismeteld and varido < 64:
            try:
                request = serviceAdmin.users().insert(body=tanar).execute()
                ismeteld=False
            except HttpError as err:
                if err.resp.status in [409]:
                    ismeteld=False
                elif err.resp.status in [403, 503]:
                    time.sleep(varido+random())
                    varido=varido*2
                else:
                    print(err)
                    sys.exit()
            print("*", end="")
        if ismeteld:
            print(err)
            sys.exit()


startTime=time.time()
print("Pedagógus felhasználók létrehozása", end ="")
batch = serviceAdmin.new_batch_http_request(callback=cb_ujTanar)
kotegmeret=30
db=0
for email, veznev, kernev, om in zip(tanarEmailek, tanarVezNevek, tanarKerNevek, tanarOMek):

    tanar = {
        "primaryEmail" : email,
        "name" : {
            "givenName" : kernev,
            "familyName" : veznev
        },
        "password" : hashlib.sha1(om.encode()).hexdigest(),
        "hashFunction" : "SHA-1",
        "changePasswordAtNextLogin" : True,
        "orgUnitPath" : "/"+teszt+pedSZEnev
    }

    request = serviceAdmin.users().insert(body=tanar)
    batch.add(request, request_id=om)
    kotegmeret-=1
    print(".", end ="")
    if kotegmeret < 1:
        batch.execute()
        kotegmeret=30
        batch = serviceAdmin.new_batch_http_request(callback=cb_ujTanar)
    db+=1

batch.execute()
print(".")
print(db, "db pedagógus felhasználó létrehozva.")
elapsedTime = time.time() - startTime
print("Ez a programrész %.3fs alatt futott le." % elapsedTime)


#----------------------------------------------------------------------------------------------------------
# 3.11. Pedagógus felhasználók csoportokba sorolása
#          - Pedagógusok csoport
#          - Tanterem_tanárok csoport
#          - osztálycsoportok
#

def cb_ujCsopTagsag(request_id, response, exception):
    if exception is not None and exception.resp.status not in [409]:
        
        email=request_id[:request_id.find("!")]
        csopAz=request_id[request_id.find("!")+1:]
        
        tag = { "email" : email,
                "role" : "MEMBER" }
        
        ismeteld=True
        varido=1

        while ismeteld and varido < 64:
            try:
                
                request=serviceAdmin.members().insert(body=tag, groupKey=csopAz).execute()
                ismeteld=False
            except HttpError as err:
                if err.resp.status in [409]:
                    ismeteld=False
                elif err.resp.status in [403, 503]:
                    time.sleep(varido+random())
                    varido=varido*2
                else:
                    print(err)
                    sys.exit()
            print("*", end="")
        if ismeteld:
            print(err)
            sys.exit()

startTime=time.time()
print("Pedagógus felhasználók csoportokba sorolása", end ="")
batch = serviceAdmin.new_batch_http_request(callback=cb_ujCsopTagsag)
kotegmeret=30
db=0
for email, oszt in zip(tanarEmailek, tanarOsztok):

    tag = { "email" : email,
            "role" : "MEMBER" }

    csopAz=teszt+"ped@"+suliDomain
    request=serviceAdmin.members().insert(body=tag, groupKey=csopAz)
    batch.add(request, request_id=email+"!"+csopAz)
    kotegmeret-=1

    csopAz="tanterem_tanarok@"+suliDomain
    request=serviceAdmin.members().insert(body=tag, groupKey=csopAz)
    batch.add(request, request_id=email+"!"+csopAz)
    kotegmeret-=1

    if oszt:
        csopAz=teszt+oszt+"@"+diakDomain
        request=serviceAdmin.members().insert(body=tag, groupKey=teszt+oszt+"@"+diakDomain)
        batch.add(request, request_id=email+"!"+csopAz)
        kotegmeret-=1

    print(".", end ="")
    if kotegmeret < 1:
        batch.execute()
        kotegmeret=30
        batch = serviceAdmin.new_batch_http_request(callback=cb_ujCsopTagsag)
    db+=1

batch.execute()
print(".")
print(db, "pedagógus felhasználó lett csoportokba sorolva.")
elapsedTime = time.time() - startTime
print("Ez a programrész %.3fs alatt futott le." % elapsedTime)



#----------------------------------------------------------------------------------------------------------
# 3.12. Diák felhasználók létrehozása
#        (és elhelyezése a megfelelő szervezeti egységekben)
#

def cb_ujDiak(request_id, response, exception):
    if exception is not None and exception.resp.status not in [409]:
        
        om=request_id
        diakIdx=diakOMek.index(om)
        
        email=diakEmailek[diakIdx]
        veznev=diakVezNevek[diakIdx]
        kernev=diakKerNevek[diakIdx]
        oszt==diakOsztalya[diakIdx]

        if oszt:
            diakOszt=oszt
            diakOsztEmail=oszt
        else:
            diakOszt=vendegSZEnev
            diakOsztEmail="vendeg"

        diak = { "primaryEmail" : email,
                 "name" : {
                    "givenName" : kernev,
                    "familyName" : veznev },
                 "password" : hashlib.sha1(om.encode()).hexdigest(),
                 "hashFunction" : "SHA-1",
                 "changePasswordAtNextLogin" : True,
                 "orgUnitPath" : "/"+teszt+diakokSZEnev+"/"+tanev+"/"+diakOszt }

        ismeteld=True
        varido=1

        while ismeteld and varido < 64:
            try:
                request = serviceAdmin.users().insert(body=diak).execute()
                ismeteld=False
            except HttpError as err:
                if err.resp.status in [409]:
                    ismeteld=False
                elif err.resp.status in [403, 503]:
                    time.sleep(varido+random())
                    varido=varido*2
                else:
                    print(err)
                    sys.exit()
            print("*", end="")
        if ismeteld:
            print(err)
            sys.exit()

startTime=time.time()
print("Diák felhasználók létrehozása", end ="")
batch = serviceAdmin.new_batch_http_request(callback=cb_ujDiak)
kotetmeret=30
db=0
for email, veznev, kernev, om, oszt in zip(diakEmailek, diakVezNevek, diakKerNevek, diakOMek, diakOsztalya):

    if oszt:
        diakOszt=oszt
        diakOsztEmail=oszt
    else:
        diakOszt=vendegSZEnev
        diakOsztEmail="vendeg"

    diak = { "primaryEmail" : email,
             "name" : {
                "givenName" : kernev,
                "familyName" : veznev },
             "password" : hashlib.sha1(om.encode()).hexdigest(),
             "hashFunction" : "SHA-1",
             "changePasswordAtNextLogin" : True,
             "orgUnitPath" : "/"+teszt+diakokSZEnev+"/"+tanev+"/"+diakOszt }
    
    request = serviceAdmin.users().insert(body=diak)
    batch.add(request, request_id=om)
    db+=1
    kotetmeret-=1
    print(".", end ="")
    if kotetmeret < 1:
        batch.execute()
        kotetmeret=30
        batch = serviceAdmin.new_batch_http_request(callback=cb_ujDiak)

batch.execute()
print(".")
print(db, " diák felhasználó lett létrehozva.")
elapsedTime = time.time() - startTime
print("Ez a programrész %.3fs alatt futott le." % elapsedTime)



#
# Diák felhasználók csoportokba sorolása
#

def cb_diakCsopPlusz(request_id, response, exception):
    if exception is not None and exception.resp.status not in [409]:

        email=request_id
        oszt=diakOsztalya[diakEmailek.index(email)]

        if oszt:
            diakOszt=oszt
            diakOsztEmail=oszt
        else:
            diakOszt=vendegSZEnev
            diakOsztEmail="vendeg"

        tag = { "email" : email,
                "role" : "MEMBER" }

        ismeteld=True
        varido=1

        while ismeteld and varido < 64:
            try:
                request = serviceAdmin.members().insert(body=tag, groupKey=teszt+diakOsztEmail+"@"+diakDomain).execute()
                ismeteld=False
            except HttpError as err:
                if err.resp.status in [409]:
                    ismeteld=False
                elif err.resp.status in [403, 503]:
                    time.sleep(varido+random())
                    varido=varido*2
                else:
                    print(err)
                    sys.exit()
            print("*", end="")
        if ismeteld:
            print(err)
            sys.exit()

startTime=time.time()
print("Diák felhasználók csoportokba sorolása", end ="")
batch = serviceAdmin.new_batch_http_request(callback=cb_diakCsopPlusz)
kotegmeret=30
db=0
for email, oszt in zip(diakEmailek, diakOsztalya):

    if oszt:
        diakOszt=oszt
        diakOsztEmail=oszt
    else:
        diakOszt=vendegSZEnev
        diakOsztEmail="vendeg"

    tag = { "email" : email,
            "role" : "MEMBER" }

    request=serviceAdmin.members().insert(body=tag, groupKey=teszt+diakOsztEmail+"@"+diakDomain)
    batch.add(request, request_id=email)
    kotegmeret-=1
    print(".", end ="")
    if kotegmeret < 1:
        batch.execute()
        kotegmeret=30
        batch = serviceAdmin.new_batch_http_request(callback=cb_diakCsopPlusz)
    db+=1

batch.execute()
print(".")
print(db, "db diák felhasználó lett csoportokba sorolva.")
elapsedTime = time.time() - startTime
print("Ez a programrész %.3fs alatt futott le." % elapsedTime)



#----------------------------------------------------------------------------------------------------------
# 3.13. Kurzusok (osztálytermek) létrehozása
#

def cb_ujKurzus(request_id, response, exception):
    if exception is not None and exception.resp.status not in [409]:
        
        kurzusAlias=request_id
        kurzusIdx=kurzusAliasok.index(kurzusAlias)
        
        nev=kurzusNevek[kurzusIdx]
        csop=kurzusCsopok[kurzusIdx]
        tant=kurzusTanok[kurzusIdx]
        OM=kurzusTanarOMek[kurzusIdx]
        
        i=tanarOMek.index(OM[:11])
        tanarNev=tanarVezNevek[i]+" "+tanarKerNevek[i]
        tanarEmail=tanarEmailek[i]

        kurzus = { "id" : kurzusAlias,
                   "name": nev,
                   "descriptionHeading": "Üdvözöllek a "+csop+" csoport "+tant+" kurzusán!",
                   "description": "A "+csop+" csoport "+tant+" kurzusa, melyet "+tanarNev+" tart.",
                   "ownerId": tanarEmail,
                   "courseState": "ACTIVE" }

        ismeteld=True
        varido=1

        while ismeteld and varido < 64:
            try:
                request = serviceClassroom.courses().create(body=kurzus).execute()
                ismeteld=False
            except HttpError as err:
                if err.resp.status in [409]:
                    ismeteld=False
                elif err.resp.status in [403, 503]:
                    time.sleep(varido+random())
                    varido=varido*2
                else:
                    print(err)
                    sys.exit()
            print("*", end="")
        if ismeteld:
            print(err)
            sys.exit()


startTime=time.time()
print("A", len(kurzusNevek), "db kurzus létrehozása folyamatban", end ="")
batch = serviceClassroom.new_batch_http_request(callback=cb_ujKurzus)
k=30
db=0
for nev, csop, tant, OM, kurzusAlias  in zip(kurzusNevek, kurzusCsopok, kurzusTanok, kurzusTanarOMek, kurzusAliasok):

    i=tanarOMek.index(OM[:11])
    tanarNev=tanarVezNevek[i]+" "+tanarKerNevek[i]
    tanarEmail=tanarEmailek[i]

    kurzus = {
        "id" : kurzusAlias,
        "name": nev,
        "descriptionHeading": "Üdvözöllek a "+csop+" csoport "+tant+" kurzusán!",
        "description": "A "+csop+" csoport "+tant+" kurzusa, melyet "+tanarNev+" tart.",
        "ownerId": tanarEmail,
        "courseState": "ACTIVE" }

    request = serviceClassroom.courses().create(body=kurzus)
    batch.add(request, request_id=kurzusAlias)
    k-=1
    if k < 1:
        batch.execute()
        k=30
        batch = serviceClassroom.new_batch_http_request(callback=cb_ujKurzus)

    db+=1
    print(".", end ="")

batch.execute()
print(".")
print(db, "db kurzus került létrehozásra.")
elapsedTime = time.time() - startTime
print("Ez a programrész %.3fs alatt futott le." % elapsedTime)



#----------------------------------------------------------------------------------------------------------
# 3.14. További tanárok hozzárendelése a kurzusokhoz
#

def cb_tanarPlusz(request_id, response, exception):
    if exception is not None and exception.resp.status not in [409]:

        kurzusAlias=request_id[:request_id.find("/")]
        tanarEmail=request_id[request_id.find("/")+1:]
        tanar = { "userId" : tanarEmail }

        ismeteld=True
        varido=1

        while ismeteld and varido < 64:
            try:
                request = serviceClassroom.courses().teachers().create(courseId=kurzusAlias,  body=tanar).execute()
                ismeteld=False
            except HttpError as err:
                if err.resp.status in [409]:
                    ismeteld=False
                elif err.resp.status in [403, 503]:
                    time.sleep(varido+random())
                    varido=varido*2
                else:
                    print(err)
                    sys.exit()
            print("*", end="")
        if ismeteld:
            print(err)
            sys.exit()

startTime=time.time()
print("További tanárok hozzáadása a kurzusokhoz", end ="")
batch = serviceClassroom.new_batch_http_request(callback=cb_tanarPlusz)
k=30
db=0
for kurzusAlias, OM  in zip(kurzusAliasok, kurzusTanarOMek):

    if len(OM)>11:
        s=OM[12:]
        while len(s)>0:

            tanarEmail=tanarEmailek[tanarOMek.index(s[:11])]
            tanar = { "userId": tanarEmail }

            tanar = serviceClassroom.courses().teachers().create(courseId=kurzusAlias,  body=tanar)
            batch.add(tanar, request_id=kurzusAlias+"/"+tanarEmail)
            s=s[12:]
            db+=1
            k-=1
            if k < 1:
                batch.execute()
                k=30
                batch = serviceClassroom.new_batch_http_request(callback=cb_tanarPlusz)
    print(".", end ="")


batch.execute()
print(".")
print(db, "módosítás került adminisztrálásra.")
elapsedTime = time.time() - startTime
print("Ez a programrész %.3fs alatt futott le." % elapsedTime)



#==========================================================================================================
# 3.15. Diákok beiskolázása (hozzárendelése a kurzusokhoz)
#

def cb_diakPlusz(request_id, response, exception):
    if exception is not None and exception.resp.status not in [409]:

        alias=request_id[:request_id.find("/")]
        diakEmail=request_id[request_id.find("/")+1:request_id.find(".hu")+3]
        diak = { "userId" : diakEmail }

        ismeteld=True
        varido=1

        while ismeteld and varido < 64:
            try:
                request = serviceClassroom.courses().students().create(courseId=alias, body=diak).execute()
                ismeteld=False
            except HttpError as err:
                if err.resp.status in [409]:
                    ismeteld=False
                elif err.resp.status in [403, 503, 429]:
                    time.sleep(varido+random())
                    varido=varido*2
                else:
                    print(err)
                    sys.exit()
            print("*", end="")
        if ismeteld:
            print(err)
            sys.exit()


startTime=time.time()
print("Diákok beiskolázása", end ="")
batch = serviceClassroom.new_batch_http_request(callback=cb_diakPlusz)
k=20
db=0
for sor in ml.rows:
    if str(sor[1].value) != "Vezetéknév" and sor[7].value is not None:
        diakEmail=diakEmailek[diakOMek.index(str(sor[3].value))]
        diak = { "userId" : diakEmail }
        alias="d:"+teszt+str(sor[4].value)+str(sor[5].value)+tanev
        alias=alias.replace(" ", "")
        alias=alias.replace(".", "")
        alias=alias.replace("-", "")
        alias=alias.replace("/", "")
        print(".", end ="")
        request = serviceClassroom.courses().students().create(courseId=alias, body=diak)
        batch.add(request, request_id=alias+"/"+diakEmail+str(k))
        k-=1
        if k < 1 :
            batch.execute()
            k=20
            batch = serviceClassroom.new_batch_http_request(callback=cb_diakPlusz)

    db+=1

batch.execute()
print("")
print(db, "db jelentkezés történt a kurzusokra.")
elapsedTime = time.time() - startTime
print("Ez a programrész %.3fs alatt futott le." % elapsedTime)

ido = (time.time() - totalTime)
perc=ido//60
masodperc=ido%60
print("Teljes futásidő: {:0.0f}:{:0.0f}".format(perc, masodperc))

tbl.close()
ped1TBL.close()
ped2TBL.close()
